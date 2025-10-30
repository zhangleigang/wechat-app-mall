from openpyxl import load_workbook
import json

# 加载Excel文件
wb = load_workbook('./doc/question and answer.xlsx')
ws = wb.active

# 获取数据
all_data = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] and row[1]:  # 确保分类和问题不为空
        all_data.append({
            'category': row[0],
            'question': row[1],
            'answer': row[2] if row[2] else ''
        })

# 按分类分组
category_data = {}
for item in all_data:
    if item['category'] not in category_data:
        category_data[item['category']] = []
    category_data[item['category']].append(item)

# 生成categories数组
categories = []
category_keys = {
    'HDFS': 'hdfs',
    'MapReduce': 'mr',
    'YARN': 'yarn',
    'Kafka': 'kafka',
    'HBase': 'hbase',
    'Hive': 'hive',
    'Spark': 'spark',
    'Flink': 'flink',
    '数据倾斜': 'skew'
}

for category_name in category_data.keys():
    key = category_keys.get(category_name, category_name.lower().replace(' ', '-'))
    categories.append({"key": key, "name": category_name})

# 生成topics数组
topics = []
topic_id_counter = 1

for category_name, items in category_data.items():
    category_key = category_keys.get(category_name, category_name.lower().replace(' ', '-'))
    
    # 为每个分类创建多个主题，每个主题包含一些问题
    questions_per_topic = 3
    for i in range(0, len(items), questions_per_topic):
        topic_questions = items[i:i+questions_per_topic]
        
        # 生成主题标题和摘要
        title = f"{category_name} 常见问题 {i//questions_per_topic + 1}"
        summary = topic_questions[0]['question'][:50] + '...' if len(topic_questions[0]['question']) > 50 else topic_questions[0]['question']
        
        # 提取常见问题
        faqs = [q['question'] for q in topic_questions[:2]]  # 每个主题最多2个FAQ
        
        # 创建主题
        topic = {
            "id": f"{category_key}-topic-{topic_id_counter}",
            "title": title,
            "summary": summary,
            "tags": [category_name, "面试"],
            "categoryKey": category_key,
            "faqs": faqs,
            "pitfalls": ["准备不充分", "回答不全面"],
            "examples": ["完整回答流程", "结合实际项目经验"]
        }
        topics.append(topic)
        topic_id_counter += 1

# 生成JavaScript代码
js_code = f"""const categories = {json.dumps(categories, ensure_ascii=False, indent=2)}

const topics = {json.dumps(topics, ensure_ascii=False, indent=2)}

module.exports = {
  categories,
  topics
}
"""

# 保存到文件
with open('updated_knowledge.js', 'w', encoding='utf-8') as f:
    f.write(js_code)

print(f"成功生成updated_knowledge.js文件")
print(f"生成了 {len(categories)} 个分类")
print(f"生成了 {len(topics)} 个主题")
print("\n分类列表:")
for cat in categories:
    print(f"  - {cat['name']} ({cat['key']})")