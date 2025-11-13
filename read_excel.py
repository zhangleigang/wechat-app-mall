from openpyxl import load_workbook

# 加载Excel文件
wb = load_workbook('./doc/question and answer_with_answers.xlsx')

# 获取活动工作表
ws = wb.active

# 打印基本信息
print('Sheet title:', ws.title)
print('Total rows:', ws.max_row)
print('Total columns:', ws.max_column)

# 打印前10行数据
print('\nFirst 10 rows of data:')
for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
    print(row)

# 获取表头
headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
print('\nHeaders:', headers)

# 尝试读取更多数据以了解结构
print('\nRows 2-5 with column names:')
for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=5, values_only=True), 2):
    print(f'Row {row_idx}:')
    for col_idx, (header, value) in enumerate(zip(headers, row)):
        print(f'  {header}: {value}')